import logging
import time
from src.const.global_map import RESOURCE_MAP
from tempfile import NamedTemporaryFile
from fastapi.templating import Jinja2Templates
from fastapi import Request
from fastapi import  WebSocket, WebSocketDisconnect
from datetime import datetime, timedelta
from typing import List
from src.utils.basemodel import app_schemas as Schema 
import json
from tqdm import tqdm
from fastapi import Request,File, UploadFile,Form
from time import sleep
from minio import Minio
from minio.error import S3Error
import os
import io
import pandas as pd
from kafka import KafkaConsumer, consumer
import msoffcrypto
import io
import pandas as pd
import src.const.const_map as CONST_MAP
from openpyxl import Workbook,load_workbook

app_logger = logging.getLogger("app_logger")

app = RESOURCE_MAP["fastapi_app"]

# src_path = os.getcwd()


def s3_cred(credential_path: str):
    with open(credential_path, "r") as f:
        return json.load(f)
  
    
def minio_upload_file(content: io.BytesIO, s3_key: str):
    s3_credential = s3_cred(os.path.join(os.getcwd(), "resource/minio_credential.json"))
    try:
        client = Minio(
            s3_credential["MINIO_END_POINT"],
            access_key=s3_credential["MINIO_ACCESS_KEY_ID"],
            secret_key=s3_credential["MINIO_SECRET_ACCESS_KEY"],
            secure=True,
        )
        content.seek(0)
        client.put_object(
            s3_credential["MINIO_STORAGE_BUCKET_NAME"],
            object_name=s3_key,
            data=content,
            length=-1,
            part_size=10 * 1024 * 1024,
        )
        
        url =client.get_presigned_url(
            "GET",
            s3_credential["MINIO_STORAGE_BUCKET_NAME"],
            s3_key,
            expires=timedelta(hours=2),
        )
        return url
    except Exception as e:
        raise RuntimeError('Upload error for key "%s".' % s3_key) from e

    
def minio_download_to_bytes(s3_key: str) -> bytes:
    s3_credential = s3_cred(os.path.join(os.getcwd(), "resource/minio_credential.json"))
    try:
        client = Minio(
            s3_credential["MINIO_END_POINT"],
            access_key=s3_credential["MINIO_ACCESS_KEY_ID"],
            secret_key=s3_credential["MINIO_SECRET_ACCESS_KEY"],
            secure=True,
        )
        response = client.get_object(
            bucket_name=s3_credential["MINIO_STORAGE_BUCKET_NAME"], object_name=s3_key
        )
        try:
            bytearr = io.BytesIO(response.data).read()
            return bytearr
        except Exception as e:
            raise RuntimeError('Other error when read from key "%s".' % s3_key) from e
        finally:
            response.close()
            response.release_conn()

    except S3Error as e:
        if e.code == "NoSuchKey":
            raise RuntimeError('The input key "%s" does not exist.' % s3_key) from e
        else:
            raise RuntimeError('Other download error for key "%s": %s' % (s3_key, e)) from e
    except Exception as e:
        raise RuntimeError('Other error for key "%s".' % s3_key) from e

class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            await connection.send_text(message)


manager = ConnectionManager()

@app.websocket("/ws/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    await manager.connect(websocket)
    now = datetime.now()
    current_time = now.strftime("%H:%M")
    try:
        while True:
            data = await websocket.receive_text()
            # await manager.send_personal_message(f"You wrote: {data}", websocket)
            message = {"time":current_time,"clientId":client_id,"message":data}

            await manager.broadcast(json.dumps(message))
         
    except KeyboardInterrupt:
            print("Aborted by user...")
    finally:
            consumer.close()
            manager.disconnect(websocket)
            message = {"time":current_time,"clientId":client_id,"message":"Offline"}
            await manager.broadcast(json.dumps(message))
        
        
@app.post("/upload/{token_id}/")
async def upload( token_id :str,file: UploadFile):
    print("UPLOAD FILE.....")
    try:
        url = minio_upload_file(file.file,f"{token_id}.xlsb")
        # with open(f"tmp/{token_id}.xlsb","wb") as f:
        #     content =  await file.read()
        #     f.write(content)
        # print(url)
    except Exception as e:
        return {"success": False}
    return {"success": True}


@app.post("/file_handling")
async def file_handling(input_map : Schema.UUIDSchemas, request: Request):
    uuid_token = input_map.uuid_token
    try:
        url=handle_excel(uuid_token=uuid_token)
    except Exception as ex:
        print(ex)
        return {"content": "error format file"}
    return {"content":url}


@app.post("/done_task")
async def file_handling(input_map : Schema.DoneTaskSchemas, request: Request):
    now = datetime.now()
    current_time = now.strftime("%H:%M")
    message = {"time":current_time,"clientId":input_map.uuid_token,"uuid_token":input_map.uuid_token, "url_download": input_map.url}
    print(message)
    await manager.broadcast(json.dumps(message))
    return {"success":False}


def handle_excel(uuid_token):
    start_time = time.time()
    print("Download Format File")
    with open("log.txt","w") as file:
        format_file = minio_download_to_bytes("format_output.xlsx")
        st_open = time.time()
        wb = load_workbook(io.BytesIO(format_file))
        file.write(f"Load format output {time.time()-st_open}s")
        byte_excel = minio_download_to_bytes(uuid_token+".xlsb")
        print("Done Download Format File")
        temp = io.BytesIO()
        password = 'KPMG@2023'
        excel = msoffcrypto.OfficeFile(io.BytesIO(byte_excel))
        excel.load_key(password)
        excel.decrypt(temp)
        st_input = time.time()
        df_input = pd.read_excel(temp,sheet_name=CONST_MAP.sheet_name_input)
        file.write(f"Load file input {time.time()-st_input}s")
        df_output = df_input
        
        df_output["1. CUSTOMER"]["CPTY_TYPE_1"] = df_output["1. CUSTOMER"].index.map(lambda x: f"=VLOOKUP(B{x+2}&C{x+2},'9. REG TABLE CAL'!B:D,2,0)" )
        df_output["1. CUSTOMER"]["CPTY_SUB_TYPE_1"] = df_output["1. CUSTOMER"].index.map(lambda x: f"=IFERROR(VLOOKUP(B{x+2}&C{x+2},'9. REG TABLE CAL'!B:D,3,0),\"Check\")" )
        df_output["1. CUSTOMER"]["CUST_RATING_CD"] = df_output["1. CUSTOMER"].index.map(lambda x: 
            f'=IF(AND(E{x+2}&D{x+2}<>"NANA",E{x+2}&D{x+2}<>""),VLOOKUP(E{x+2}&D{x+2},\'9. REG TABLE CAL\'!N:O,2,0),IF(AND(OR(E{x+2}&D{x+2}="NANA",E{x+2}&D{x+2}=""),B{x+2}<>"FIN_INST"),"LT7",IF(AND(OR(E{x+2}&D{x+2}="NANA",E{x+2}&D{x+2}=""),B{x+2}="FIN_INST"),IF(VLOOKUP(A{x+2},SCRA!$A:$A,1,FALSE)=\'1. CUSTOMER\'!A{x+2},VLOOKUP(A{x+2},SCRA!$A:$J,10,FALSE),"NA"))))')

        df_output["2. SCRA"]["SCRA Group"]= df_output["2. SCRA"].index.map(lambda x: None)
        df_output["2. SCRA"]["SCRA Group"][0]="SCRA Group"
        df_output["2. SCRA"]["SCRA Group"][1:] =  df_output["2. SCRA"].index[1:].map(lambda x: f'=IF(OR(B{x+2}="",C{x+2}="",D{x+2}="",E{x+2}="",F{x+2}="",G{x+2}="",H{x+2}="",I{x+2}=""),"INVALID",IF( AND(B{x+2}="Y",C{x+2}="Y",I{x+2}="Y"),"Group A and satisfy #",IF(AND(B{x+2}="Y",C{x+2}="Y"),"Group A",IF(E{x+2}="Y","Group B","Group C"))))')

        df_output["4. COLLATERAL"]["COLL_ORIGINAL_MATURITY"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=(I{x+2}-H{x+2})/365")
        df_output["4. COLLATERAL"]["COLL_REMAINING_MATURITY"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=(I{x+2}-J{x+2})/365")
        df_output["4. COLLATERAL"]["COLL_ORIGINAL_MATURITY_BUCKET"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(L{x+2}<='7. REG TABLE'!$U$9,'7. REG TABLE'!$V$9,IF(L{x+2}<='7. REG TABLE'!$U$10,'7. REG TABLE'!$V$10,IF(L{x+2}<='7. REG TABLE'!$U$11,'7. REG TABLE'!$V$11,IF(L{x+2}<='7. REG TABLE'!$U$12,'7. REG TABLE'!$V$12,IF(L{x+2}<='7. REG TABLE'!$U$13,'7. REG TABLE'!$V$13,\"NA\")))))")
        df_output["4. COLLATERAL"]["COLL_REMAINING_MATURITY_BUCKET"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(M{x+2}<='7. REG TABLE'!$U$9,'7. REG TABLE'!$V$9,IF(M{x+2}<='7. REG TABLE'!$U$10,'7. REG TABLE'!$V$10,IF(M{x+2}<='7. REG TABLE'!$U$11,'7. REG TABLE'!$V$11,IF(M{x+2}<='7. REG TABLE'!$U$12,'7. REG TABLE'!$V$12,IF(M{x+2}<='7. REG TABLE'!$U$13,'7. REG TABLE'!$V$13,\"NA\")))))")
        df_output["4. COLLATERAL"]["COLL_RATING_CD"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(K{x+2}=\"\",\"LT7\",VLOOKUP(K{x+2},'7. REG TABLE'!D:E,2,0))")
        df_output["4. COLLATERAL"]["CRM_CD"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(OR(D{x+2}=\"CRM_DEBT_OTH_BANK\",D{x+2}=\"CRM_DEBT_OTH_CORP\",D{x+2}=\"CRM_VAL_PAPER_FOR_GOV\",D{x+2}=\"CRM_VAL_PAPER_GOV_SBV\",D{x+2}=\"CRM_DEBT_PSE\"),VLOOKUP(D{x+2}&P{x+2}&N{x+2},'9. REG TABLE CAL'!J:L,2,FALSE),VLOOKUP(D{x+2}&P{x+2},'9. REG TABLE CAL'!J:L,2,FALSE))")
        df_output["4. COLLATERAL"]["ELIGIBLE_CRM"] = df_output["4. COLLATERAL"].index.map(lambda x:f"=IF(OR(D{x+2}=\"CRM_DEBT_OTH_BANK\",D{x+2}=\"CRM_DEBT_OTH_CORP\",D{x+2}=\"CRM_VAL_PAPER_FOR_GOV\",D{x+2}=\"CRM_VAL_PAPER_GOV_SBV\",D{x+2}=\"CRM_DEBT_PSE\"),VLOOKUP(D{x+2}&P{x+2}&N{x+2},'9. REG TABLE CAL'!J:L,3,FALSE),VLOOKUP(D{x+2}&P{x+2},'9. REG TABLE CAL'!J:L,3,FALSE))")
        df_output["4. COLLATERAL"]["HAIRCUT_CD"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(R{x+2}=\"NO\",\"N/A\",INDEX('9. REG TABLE CAL'!$T$10:$U$190,MATCH(IF(OR(D{x+2}=\"CRM_DEBT_OTH_BANK\",D{x+2}=\"CRM_DEBT_OTH_CORP\",D{x+2}=\"CRM_DEBT_PSE\",D{x+2}=\"CRM_VAL_PAPER_FOR_GOV\",D{x+2}=\"CRM_VAL_PAPER_GOV_SBV\"),TRIM(CONCATENATE(D{x+2},Q{x+2},O{x+2},P{x+2})),TRIM(CONCATENATE(D{x+2},Q{x+2}))),'9. REG TABLE CAL'!$T$10:$T$190,0),2))")
        df_output["4. COLLATERAL"]["HAIRCUT%"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(S{x+2}=\"N/A\",100%,VLOOKUP(S{x+2},'7. REG TABLE'!$M$8:$N$22,2,0))")
        df_output["4. COLLATERAL"]["COLL_CRM_ELIGIBLE %"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=IF(R{x+2}=\"NO\",0,IF(ISERROR(F{x+2}/SUMIFS(F:F,B:B,B{x+2},R:R,\"YES\")),0,F{x+2}/SUMIFS(F:F,B:B,B{x+2},R:R,\"YES\")))")
        df_output["4. COLLATERAL"]["ALLOCATED_COLLATERAL_AFTER_MT_MISMATCH_AND_HAIRCUT"] = df_output["4. COLLATERAL"].index.map(lambda x: f"=SUMIFS('6. EXPOSURE'!AQ:AQ,'6. EXPOSURE'!B:B,'4. COLLATERAL'!B{x+2})*'4. COLLATERAL'!U{x+2}*(1-T{x+2})")

        df_output["5. GUARANTEE"]["GUARANTEE_ORIGINAL_MATURITY"]=df_output["5. GUARANTEE"].index.map(lambda x: f"=(I{x+2}-H{x+2})/365")
        df_output["5. GUARANTEE"]["GUARANTEE_REMAINING_MATURITY"]=df_output["5. GUARANTEE"].index.map(lambda x: f"=(I{x+2}-J{x+2})/365")
        df_output["5. GUARANTEE"]["GUARANTEE_CRM_ELIGIBLE %"]=df_output["5. GUARANTEE"].index.map(lambda x: f"=IF(D{x+2}=\"N\",0,IF(ISERROR(E{x+2}/SUMIFS(E:E,B:B,B{x+2},D:D,\"Y\")),0,E{x+2}/SUMIFS(E:E,B:B,B{x+2},D:D,\"Y\")))")
        df_output["5. GUARANTEE"]["ALLOCATED_GUARANTEE_AFTER_MT_MISMATCH"]=df_output["5. GUARANTEE"].index.map(lambda x: f"=SUMIFS('6. EXPOSURE'!AT:AT,'6. EXPOSURE'!B:B,'5. GUARANTEE'!B{x+2},'6. EXPOSURE'!AO:AO,\">=\"&K{x+2})*N{x+2}")
        df_output["5. GUARANTEE"]["GUARANTEE_RWA"]=df_output["5. GUARANTEE"].index.map(lambda x: f"=SUMIFS('6. EXPOSURE'!AU:AU,'6. EXPOSURE'!B:B,'5. GUARANTEE'!B{x+2},'6. EXPOSURE'!AO:AO,\">=\"&K{x+2})*N{x+2}*K{x+2}")

        df_output["OD"]["Transactor flag"] = df_output["OD"].index.map(lambda x: f"=IFERROR(IF(AND(COUNT(D{x+2}:O{x+2})=12,SUM(IF((E{x+2}:O{x+2}/D{x+2}:N{x+2})<1,0,1))=11),\"Y\",\"N\"),\"N\")")
        df_output["CC"]["Transactor flag"] = df_output["CC"].index.map(lambda x: f'=IF(AND(COUNT(E{x+2}:AB{x+2})={x+2}4,SUM(IF(E{x+2}:P{x+2}-Q{x+2}:AB{x+2}<0,0,1))=0),"Y","N")')
        print("Copy Excel")
        for sheetname  in CONST_MAP.sheet_name_input[:7]:
            for index, row in df_output[sheetname].iterrows():
                for col_num, value in enumerate(row, 1):
                    wb[sheetname].cell(row=index+2, column=col_num).value = value

        print("Done copy")
        st_out = time.time()
        output = io.BytesIO()
        wb.save(output)
        file.write(f"Save file {time.time()-st_out}s")
        
        url = minio_upload_file(output,f"{uuid_token}.xlsx")  
        print(url)
        del temp
        print(f"Time excute: {time.time()-start_time}s")
    return url